import openpyxl

def get_excel_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    data =[]

    for row in range(2, sheet.max_row+1):
        username = sheet.cell(row = row, column=1).value
        password = sheet.cell(row = row, column=2).value
        if username is not None:
            data.append([username,password])
            break

    return data



































